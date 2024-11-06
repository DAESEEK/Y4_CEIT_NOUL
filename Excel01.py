# xlwings  openpyxl ***  xlrd & xlwt  xlsxwriter  pyWin32 pandas **
from openpyxl import load_workbook
load_wb = load_workbook('iATERlist.xlsx',data_only=True)
load_ws = load_wb['data']
print(load_ws.cell(2,2).value)

 
#from openpyxl import
#wb = openpyxl.load_workbook('iAterlist.xlsx')
#wb_p=wb.sheetnames
#ws = wb['data']
#print(ws['A1'].value)
#print(ws['D3'].value)
#print(ws.cell(row=1,column=3).value)
#multi_cells = ws['A1':'E8']
#print(multi_cells)
#for row in multi_cells:
#    print(row[0].value,row[1].value,row[2].value,row[3].value)

#for row in ws.rows:
#     print([col.value for col in row])





